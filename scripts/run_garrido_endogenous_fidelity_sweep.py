#!/usr/bin/env python3
"""Endogenous fidelity sweep: our proxy model vs Garrido's workbook ReT, CF1-CF20.

For each thesis configuration CF1..CF20 this runs the FROZEN garrido_proxy_v1
physics with the config's risk design (design_spec_for_cfi), generating orders
ENDOGENOUSLY (not replaying Garrido's tape), and compares the workbook-visible
ReT (raw + clipped) and visible-order population against the ground-truth means
read directly from Garrido's Raw_data1/Raw_data2 workbooks.

This is a reproducibility artifact (the earlier claim that a sweep was "running"
was false at the time; this makes it a real, recomputable gate). It does NOT
open virgin tapes and trains no PPO. Primary purpose: quantify R1r vs R2r
distributional fidelity under ret_excel_visible_v1 BEFORE DRA-1.
"""
from __future__ import annotations

import argparse
import csv
import json
import statistics
from pathlib import Path
import sys

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from supply_chain.supply_chain import MFSCSimulation  # noqa: E402
from supply_chain.thesis_design import design_spec_for_cfi  # noqa: E402
from supply_chain.ret_thesis import (  # noqa: E402
    compute_order_level_ret_excel_visible_ledger as visible_ret,
)

PROXY_PATH = Path("supply_chain/data/garrido_proxy_v1_freeze_2026-07-10.json")
WB1 = Path("/Users/thom/Downloads/Raw_data1+Re.xlsx")
WB2 = Path("/Users/thom/Downloads/Raw_data2+Re.xlsx")
DEFAULT_OUTPUT = Path("results/program_d/garrido_endogenous_fidelity")


def _find_col(ws, name: str):
    for hr in (1, 2):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=hr, column=c).value
            if v is not None and str(v).strip().lower() == name.lower():
                return c, hr
    return None, None


def garrido_workbook_ret(path: Path, sheet: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    rc, hr = _find_col(ws, "ReT")
    vals = []
    for r in range(hr + 1, ws.max_row + 1):
        try:
            vals.append(float(ws.cell(row=r, column=rc).value))
        except (TypeError, ValueError):
            pass
    wb.close()
    return {
        "mean_raw": statistics.mean(vals),
        "mean_clipped": sum(min(v, 1.0) for v in vals) / len(vals),
        "n_visible": len(vals),
    }


def our_endogenous_ret(cfi: int, seeds: list[int], base: dict) -> list[dict]:
    spec = design_spec_for_cfi(cfi)
    rows = []
    for s in seeds:
        sim = MFSCSimulation(
            shifts=spec.shifts,
            initial_buffers={},
            seed=s,
            horizon=spec.horizon_hours,
            risks_enabled=True,
            enabled_risks=set(spec.enabled_risks),
            risk_overrides=dict(spec.risk_overrides),
            strict_exogenous_crn=True,
            **dict(base),
        )
        sim.run()
        v = visible_ret(list(sim.orders), current_time=sim.env.now)
        rv = list(v["ret_values"])
        rows.append(
            {
                "mean_raw": float(v["mean_ret_excel"]),
                "mean_clipped": sum(min(x, 1.0) for x in rv) / max(len(rv), 1),
                "n_visible": int(v["n_visible_rows"]),
                "n_generated": int(v["n_generated_orders"]),
                "lost": int(sim.total_unattended_orders),
            }
        )
    return rows


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--seeds", type=int, nargs="+", default=[101, 102, 103])
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)

    proxy = json.loads(PROXY_PATH.read_text(encoding="utf-8"))
    base = dict(proxy["sim_kwargs"])
    base.pop("risk_level", None)
    base.pop("seed_stream_mode", None)

    rows = []
    for cfi in range(1, 21):
        fam = "R1r" if cfi <= 10 else "R2r"
        wb = WB1 if cfi <= 10 else WB2
        g = garrido_workbook_ret(wb, f"CF{cfi}")
        ours = our_endogenous_ret(cfi, args.seeds, base)
        our_raw = statistics.mean(o["mean_raw"] for o in ours)
        our_clip = statistics.mean(o["mean_clipped"] for o in ours)
        our_n = statistics.mean(o["n_visible"] for o in ours)
        rel_gap = (our_raw - g["mean_raw"]) / g["mean_raw"] * 100 if g["mean_raw"] else float("nan")
        rows.append(
            {
                "cfi": cfi,
                "family": fam,
                "garrido_ret_raw": round(g["mean_raw"], 6),
                "garrido_ret_clipped": round(g["mean_clipped"], 6),
                "garrido_n_visible": g["n_visible"],
                "our_ret_raw": round(our_raw, 6),
                "our_ret_clipped": round(our_clip, 6),
                "our_n_visible": round(our_n, 1),
                "ret_rel_gap_pct": round(rel_gap, 1),
                "seed_spread_raw": round(
                    max(o["mean_raw"] for o in ours) - min(o["mean_raw"] for o in ours), 6
                ),
                "n_seeds": len(args.seeds),
            }
        )
        print(f"[fidelity] CF{cfi} {fam} garrido={g['mean_raw']:.5f} ours={our_raw:.5f} gap={rel_gap:+.1f}%", flush=True)

    with (args.output_dir / "fidelity_by_config.csv").open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)

    def fam_stats(fam):
        gaps = [r["ret_rel_gap_pct"] for r in rows if r["family"] == fam]
        return {
            "mean_rel_gap_pct": round(statistics.mean(gaps), 1),
            "min_rel_gap_pct": min(gaps),
            "max_rel_gap_pct": max(gaps),
            "abs_mean_rel_gap_pct": round(statistics.mean(abs(x) for x in gaps), 1),
        }

    summary = {
        "kind": "garrido_endogenous_fidelity_sweep",
        "metric": "ret_excel_visible_v1",
        "seeds": args.seeds,
        "R1r": fam_stats("R1r"),
        "R2r": fam_stats("R2r"),
        "interpretation": (
            "R1r reasonable on average with per-config scatter; R2r is NOT tightly "
            "calibrated (scattered, mean-biased) — the family DRA-1 depends on "
            "(R22/R23). Fidelity of R2r mechanics must be characterized before "
            "trusting DRA-1 spatial-allocation dynamics."
        ),
    }
    (args.output_dir / "summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
    print(json.dumps(summary, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
