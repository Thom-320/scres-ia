#!/usr/bin/env python3
"""Export OUR DES order ledger in Garrido's Raw_data Excel format for side-by-side audit.

Garrido's `Raw_data1+Re.xlsx` / `Raw_data2+Re.xlsx` are per-ORDER workbooks: one row per order
with Q, OPTj, OATj, CTj, LT, APj/RPj/DPj, the risk columns (R11-R14 / R21-R24), cumulative Bt/Ut,
and the order-level ReT (Excel formula). This script runs a config through the faithful DES and
writes the SAME per-order schema so the two workbooks can be diffed/audited directly.

One sheet per (config, regime); plus an Overview sheet with the aggregate ReT + case mix.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl

from supply_chain.supply_chain import MFSCSimulation, SIMULATION_HORIZON
from supply_chain.config import THESIS_FAITHFUL_PROTOCOL as P, INVENTORY_BUFFERS
from supply_chain.ret_thesis import (
    compute_ret_per_order_excel_formula,
    order_counts_as_backorder_for_fill_rate,
    compute_order_level_ret_excel_formula,
)

# Garrido R1 family risk columns (Raw_data1). R2 family (R21-R24) appears for CF11-20.
RISK_COLS = ["R11", "R12", "R13", "R14", "R21", "R22", "R23", "R24"]


def run_sim(shifts, period, seed, regime, horizon):
    bufs = dict(INVENTORY_BUFFERS[period]) if period else None
    sim = MFSCSimulation(
        shifts=shifts, seed=seed, horizon=horizon, risks_enabled=True, risk_level=regime,
        risk_occurrence_mode="thesis_window", year_basis=P["year_basis"],
        warmup_trigger=P["warmup_trigger"], r14_defect_mode=P["r14_defect_mode"],
        downstream_q_source="figure_6_2", raw_material_flow_mode=P["raw_material_flow_mode"],
        raw_material_order_up_to_multiplier=P["raw_material_order_up_to_multiplier"],
        initial_buffers=bufs, inventory_replenishment_period=(float(period) if period else None),
    )
    sim.run()
    return sim


def order_rows(sim):
    """One Garrido-format row per order, with cumulative Bt/Ut and ReT (Excel formula)."""
    orders = sorted(sim.orders, key=lambda o: (int(getattr(o, "j", 0) or 0),
                                               float(getattr(o, "OPTj", 0.0) or 0.0)))
    rows = []
    cum_bt = 0
    cum_ut = 0
    for idx, o in enumerate(orders, start=1):
        if bool(getattr(o, "lost", False)):
            cum_ut += 1
        elif order_counts_as_backorder_for_fill_rate(o, current_time=float(sim.env.now)):
            cum_bt += 1
        ret, case = compute_ret_per_order_excel_formula(
            o, j=idx, cumulative_backorders=cum_bt, cumulative_unattended=cum_ut)
        opt = getattr(o, "OPTj", None)
        oat = getattr(o, "OATj", None)
        lt = getattr(o, "LTj", None) or getattr(o, "lead_time", None)
        ctj = (float(oat) - float(opt)) if (oat is not None and opt is not None) else None
        risks = dict(getattr(o, "ret_risk_indicators", {}) or {})
        row = {
            "j": idx, "Q": getattr(o, "quantity", getattr(o, "Q", None)),
            "OPTj": opt, "OATj": oat, "LT": lt, "CTj": ctj,
            "APj": getattr(o, "APj", None), "RPj": getattr(o, "RPj", None),
            "DPj": getattr(o, "DPj", None),
            "lost": int(bool(getattr(o, "lost", False))),
            "backorder": int(bool(getattr(o, "backorder", False))),
            "sumBt": cum_bt, "sumUt": cum_ut,
            "ReTj": ret, "case": case,
        }
        for rc in RISK_COLS:
            row[rc] = float(risks.get(rc, 0.0))
        rows.append(row)
    return rows


def write_sheet(wb, title, rows):
    ws = wb.create_sheet(title=title[:31])
    if not rows:
        ws.append(["(no orders)"]); return
    headers = ["j", "Q", "OPTj", "OATj", "LT", "CTj", "APj", "RPj", "DPj",
               *RISK_COLS, "sumBt", "sumUt", "lost", "backorder", "ReTj", "case"]
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h) for h in headers])


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--configs", default="S1_I168,S2_I168,S3_I336",
                    help="comma list like S1_I168,S3_I1344")
    ap.add_argument("--regimes", default="current,increased,severe")
    ap.add_argument("--seed", type=int, default=1)
    ap.add_argument("--horizon", type=int, default=SIMULATION_HORIZON)
    ap.add_argument("--output", default="outputs/audits/des_garrido_format_2026-06-27.xlsx")
    args = ap.parse_args()
    regimes = [r.strip() for r in args.regimes.split(",") if r.strip()]
    configs = []
    for tok in args.configs.split(","):
        tok = tok.strip()
        if not tok:
            continue
        s, i = tok.split("_")
        configs.append((tok, int(s[1:]), int(i[1:])))
    out = Path(args.output); out.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ov = wb.active; ov.title = "Overview"
    ov.append(["DES order ledger in Garrido Raw_data format — for side-by-side audit vs "
               "Raw_data1+Re.xlsx / Raw_data2+Re.xlsx"])
    ov.append(["config", "regime", "n_orders", "mean_ReT_excel", "fill", "autotomy",
               "recovery", "risk_no_recovery", "unfulfilled"])
    for label, sh, period in configs:
        for regime in regimes:
            sim = run_sim(sh, period, args.seed, regime, args.horizon)
            rows = order_rows(sim)
            agg = compute_order_level_ret_excel_formula(sim.orders, current_time=float(sim.env.now))
            cc = agg["case_counts"]
            ov.append([label, regime, agg["n_orders"], round(agg["mean_ret_excel"], 6),
                       cc["excel_fill_rate"], cc["excel_autotomy"], cc["excel_recovery"],
                       cc["excel_risk_no_recovery"], cc["excel_unfulfilled"]])
            write_sheet(wb, f"{label}_{regime}", rows)
            print(f"  {label} {regime}: {agg['n_orders']} orders, "
                  f"mean_ReT={agg['mean_ret_excel']:.5f}", flush=True)
    wb.save(out)
    print(f"\nWROTE {out}  ({len(wb.sheetnames)-1} order-ledger sheets + Overview)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
