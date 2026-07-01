#!/usr/bin/env python3
"""Audit Garrido workbook structure, ReT branches, and DES divergence signals.

This script does not modify the original Excel files.  It treats
``Raw_data1+Re.xlsx`` and ``Raw_data2+Re.xlsx`` as the order-level ground truth
and ``Rsult_1.xlsx`` as the secondary aggregate/distribution workbook.
"""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
import statistics
import sys
from typing import Any

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from supply_chain.config import (  # noqa: E402
    INVENTORY_BUFFERS,
    THESIS_FAITHFUL_PROTOCOL as P,
    THESIS_REPLICATION_DOWNSTREAM_Q_SOURCE as DQ,
)
from supply_chain.garrido_replication import (  # noqa: E402
    DEFAULT_RAW_WORKBOOKS,
    DEFAULT_RSULT_WORKBOOK,
    audit_raw_garrido_formula,
    excel_ret_value,
    load_raw_garrido_targets,
    summarize_rsult_workbook,
)
from supply_chain.ret_thesis import compute_ret_per_order_excel_formula  # noqa: E402
from supply_chain.supply_chain import MFSCSimulation, SIMULATION_HORIZON  # noqa: E402


RAW_ROLES = {
    "Raw_data1+Re.xlsx": "Fuente primaria order-level para CF1-CF10; contiene Q, OPTj, OATj, CTj, LT, APj/RPj/DPj, riesgos R*, ReT y deltaReT por orden.",
    "Raw_data2+Re.xlsx": "Fuente primaria order-level para CF11-CF20; misma estructura que Raw_data1+Re.xlsx para la segunda familia de configuraciones.",
    "Rsult_1.xlsx": "Workbook secundario/agregado; resume distribuciones o resultados por indicador (APj, RPj, DPj, Re) y sirve para validar agregados, no para replay directo de trayectoria.",
}


DES_POLICIES = {
    "original_S1_I0": (1, 0),
    "I168_S1": (1, 168),
    "I168_S2": (2, 168),
    "I1344_S3": (3, 1344),
}


def _mean(values: list[float]) -> float:
    return float(statistics.fmean(values)) if values else 0.0


def _median(values: list[float]) -> float:
    return float(statistics.median(values)) if values else 0.0


def _p95(values: list[float]) -> float:
    if not values:
        return 0.0
    ordered = sorted(values)
    index = min(len(ordered) - 1, int(round(0.95 * (len(ordered) - 1))))
    return float(ordered[index])


def _positive_stats(values: list[float]) -> dict[str, float]:
    positive = [float(value) for value in values if float(value) > 0.0]
    return {
        "n_positive": len(positive),
        "share_positive": len(positive) / max(len(values), 1),
        "mean_positive": _mean(positive),
        "median_positive": _median(positive),
        "p95_positive": _p95(positive),
    }


def workbook_inventory(paths: list[Path]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for path in paths:
        formula_wb = load_workbook(path, data_only=False, read_only=True)
        value_wb = load_workbook(path, data_only=True, read_only=True)
        try:
            for sheet in formula_wb.sheetnames:
                ws_formula = formula_wb[sheet]
                ws_value = value_wb[sheet]
                formula_cells = 0
                first_formula = ""
                first_formula_cell = ""
                for row in ws_formula.iter_rows():
                    for cell in row:
                        value = cell.value
                        if isinstance(value, str) and value.startswith("="):
                            formula_cells += 1
                            if not first_formula:
                                first_formula = value
                                first_formula_cell = cell.coordinate
                    if first_formula and formula_cells > 100:
                        break
                rows.append(
                    {
                        "workbook": path.name,
                        "role": RAW_ROLES.get(path.name, ""),
                        "sheet": sheet,
                        "max_row": ws_value.max_row,
                        "max_column": ws_value.max_column,
                        "formula_cells_seen": formula_cells,
                        "first_formula_cell": first_formula_cell,
                        "first_formula": first_formula,
                    }
                )
        finally:
            formula_wb.close()
            value_wb.close()
    return rows


def raw_cf_rows(raw_paths: list[Path]) -> list[dict[str, Any]]:
    targets = load_raw_garrido_targets(raw_paths)
    rows: list[dict[str, Any]] = []
    for cfi, target in sorted(targets.items()):
        risks = {label: 0 for label in target.risk_columns}
        risk_active = 0
        recomputed_mismatches = 0
        max_recompute_gap = 0.0
        ap_values: list[float] = []
        rp_values: list[float] = []
        dp_values: list[float] = []
        ct_values: list[float] = []
        ret_values: list[float] = []
        ret_gt1_orders = 0
        instant_orders = 0
        late_orders = 0
        unfulfilled_orders = 0
        branch_ret: dict[str, list[float]] = {
            "excel_fill_rate": [],
            "excel_autotomy": [],
            "excel_recovery": [],
            "excel_risk_no_recovery": [],
        }
        for order in target.orders:
            active = any(value > 0.0 for value in order.risk_values.values())
            risk_active += int(active)
            for label, value in order.risk_values.items():
                risks[label] = risks.get(label, 0) + int(value > 0.0)
            ap_values.append(order.apj)
            rp_values.append(order.rpj)
            dp_values.append(order.dpj)
            ct_values.append(order.ctj)
            ret_values.append(order.ret)
            ret_gt1_orders += int(order.ret > 1.0)
            instant_orders += int(order.ctj <= 1e-9)
            late_orders += int(order.ctj > order.ltj)
            branch_ret[order.excel_case].append(order.ret)
            computed = excel_ret_value(
                j=order.j,
                lt=order.ltj,
                sum_bt=order.sum_bt,
                apj=order.apj,
                rpj=order.rpj,
                sum_ut=order.sum_ut,
                risk_values=order.risk_values.values(),
            )
            gap = abs(computed - order.ret)
            max_recompute_gap = max(max_recompute_gap, gap)
            recomputed_mismatches += int(gap > 1e-9)
        n = max(target.n_orders, 1)
        row: dict[str, Any] = {
            "cfi": cfi,
            "workbook": Path(target.source_file).name,
            "sheet": target.sheet,
            "seed": target.seed,
            "warmup_hours": target.warmup_hours,
            "horizon_hours": target.horizon_hours,
            "n_orders": target.n_orders,
            "risk_columns": "|".join(str(label) for label in target.risk_columns),
            "n_risk_columns": len(target.risk_columns),
            "ret_mean_excel": target.ret_mean_excel,
            "ret_max_excel": max(ret_values) if ret_values else 0.0,
            "ret_gt1_count": ret_gt1_orders,
            "ret_gt1_share": ret_gt1_orders / n,
            "risk_active_share": risk_active / n,
            "risk_inactive_share": 1.0 - risk_active / n,
            "instant_ct_share": instant_orders / n,
            "late_ct_share": late_orders / n,
            "unfulfilled_ct_share": unfulfilled_orders / n,
            "recomputed_mismatches": recomputed_mismatches,
            "recomputed_max_abs_gap": max_recompute_gap,
        }
        for case, count in target.case_counts_excel_formula.items():
            row[f"{case}_count"] = count
            row[f"{case}_share"] = count / n
            row[f"{case}_ret_mean"] = _mean(branch_ret[case])
        for prefix, values in (
            ("APj", ap_values),
            ("RPj", rp_values),
            ("DPj", dp_values),
            ("CTj", ct_values),
            ("ReT", ret_values),
        ):
            stats = _positive_stats(values)
            for key, value in stats.items():
                row[f"{prefix}_{key}"] = value
        for label, count in sorted(risks.items()):
            clean = str(label).replace(" ", "")
            row[f"risk_{clean}_positive_share"] = count / n
        rows.append(row)
    return rows


def rsult_rows(path: Path) -> list[dict[str, Any]]:
    summary = summarize_rsult_workbook(path)
    rows: list[dict[str, Any]] = []
    for sheet, columns in summary.get("sheets", {}).items():
        for column, stats in columns.items():
            rows.append(
                {
                    "workbook": Path(path).name,
                    "sheet": sheet,
                    "column": column,
                    **stats,
                }
            )
    return rows


def _run_des_policy(
    policy: str,
    shifts: int,
    period: int,
    seed: int,
    regime: str,
    *,
    demand_on_hand_fulfillment_delay: float,
) -> dict[str, Any]:
    buffers = dict(INVENTORY_BUFFERS[period]) if period else None
    sim = MFSCSimulation(
        shifts=shifts,
        seed=seed,
        horizon=SIMULATION_HORIZON,
        risks_enabled=True,
        risk_level=regime,
        risk_occurrence_mode="thesis_window",
        year_basis=P["year_basis"],
        warmup_trigger=P["warmup_trigger"],
        r14_defect_mode=P["r14_defect_mode"],
        downstream_q_source=DQ,
        raw_material_flow_mode=P["raw_material_flow_mode"],
        raw_material_order_up_to_multiplier=P["raw_material_order_up_to_multiplier"],
        demand_on_hand_fulfillment_delay=float(demand_on_hand_fulfillment_delay),
        initial_buffers=buffers,
        inventory_replenishment_period=(float(period) if period else None),
    )
    sim.run()
    ret = sim.compute_order_level_ret()
    ap_values: list[float] = []
    rp_values: list[float] = []
    dp_values: list[float] = []
    ct_values: list[float] = []
    risk_active = 0
    instant_orders = 0
    late_orders = 0
    unfulfilled_orders = 0
    branch_counts = {
        "excel_fill_rate": 0,
        "excel_autotomy": 0,
        "excel_recovery": 0,
        "excel_risk_no_recovery": 0,
    }
    cumulative_backorders = 0
    cumulative_unattended = 0
    for idx, order in enumerate(sorted(sim.orders, key=lambda item: int(getattr(item, "j", 0) or 0)), start=1):
        indicators = getattr(order, "ret_risk_indicators", {}) or {}
        risk_active += int(any(float(value or 0.0) > 0.0 for value in indicators.values()))
        ap_values.append(float(getattr(order, "APj", 0.0) or 0.0))
        rp_values.append(float(getattr(order, "RPj", 0.0) or 0.0))
        dp_values.append(float(getattr(order, "DPj", 0.0) or 0.0))
        lt = float(getattr(order, "LTj", 48.0) or 48.0)
        ct_raw = getattr(order, "CTj", None)
        if ct_raw is None:
            unfulfilled_orders += 1
        else:
            ct = float(ct_raw)
            ct_values.append(ct)
            instant_orders += int(ct <= 1e-9)
            late_orders += int(ct > lt)
        if bool(getattr(order, "lost", False)):
            cumulative_unattended += 1
        elif getattr(order, "OATj", None) is None:
            cumulative_backorders += 1
        _value, case = compute_ret_per_order_excel_formula(
            order,
            j=idx,
            cumulative_backorders=cumulative_backorders,
            cumulative_unattended=cumulative_unattended,
        )
        branch_counts[case] = branch_counts.get(case, 0) + 1
    n = max(len(sim.orders), 1)
    row: dict[str, Any] = {
        "policy": policy,
        "regime": regime,
        "seed": seed,
        "shifts": shifts,
        "inventory_period": period,
        "demand_on_hand_fulfillment_delay": float(demand_on_hand_fulfillment_delay),
        "n_orders": len(sim.orders),
        "mean_ret_excel_formula": float(ret["mean_ret_excel_formula"]),
        "fill_rate_order_level": float(ret["fill_rate_order_level"]),
        "risk_active_share": risk_active / n,
        "risk_inactive_share": 1.0 - risk_active / n,
        "instant_ct_share": instant_orders / n,
        "late_ct_share": late_orders / n,
        "unfulfilled_ct_share": unfulfilled_orders / n,
        "lost_orders": sum(1 for order in sim.orders if bool(getattr(order, "lost", False))),
        "pending_backorder_qty": float(getattr(sim, "pending_backorder_qty", 0.0) or 0.0),
    }
    for case, count in branch_counts.items():
        row[f"{case}_share"] = count / n
    for prefix, values in (
        ("APj", ap_values),
        ("RPj", rp_values),
        ("DPj", dp_values),
        ("CTj", ct_values),
    ):
        for key, value in _positive_stats(values).items():
            row[f"{prefix}_{key}"] = value
    return row


def des_branch_rows(
    seeds: list[int],
    regimes: list[str],
    *,
    demand_on_hand_fulfillment_delay: float,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for policy, (shifts, period) in DES_POLICIES.items():
        for regime in regimes:
            for seed in seeds:
                rows.append(
                    _run_des_policy(
                        policy,
                        shifts,
                        period,
                        seed,
                        regime,
                        demand_on_hand_fulfillment_delay=(
                            demand_on_hand_fulfillment_delay
                        ),
                    )
                )
    return rows


def aggregate_des_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for row in rows:
        grouped.setdefault((str(row["policy"]), str(row["regime"])), []).append(row)
    out: list[dict[str, Any]] = []
    metrics = [
        "mean_ret_excel_formula",
        "fill_rate_order_level",
        "risk_active_share",
        "risk_inactive_share",
        "instant_ct_share",
        "late_ct_share",
        "unfulfilled_ct_share",
        "lost_orders",
        "pending_backorder_qty",
        "excel_fill_rate_share",
        "excel_autotomy_share",
        "excel_recovery_share",
        "RPj_mean_positive",
        "RPj_median_positive",
        "RPj_p95_positive",
        "APj_mean_positive",
        "APj_median_positive",
        "APj_p95_positive",
    ]
    for (policy, regime), group in sorted(grouped.items()):
        item: dict[str, Any] = {"policy": policy, "regime": regime, "n": len(group)}
        for metric in metrics:
            item[f"{metric}_mean"] = _mean([float(row.get(metric, 0.0)) for row in group])
        out.append(item)
    return out


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    fieldnames = sorted({key for row in rows for key in row})
    with path.open("w", newline="", encoding="utf-8") as file_obj:
        writer = csv.DictWriter(file_obj, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def build_markdown(payload: dict[str, Any]) -> str:
    raw_totals = payload["raw_formula_audit"]
    lines = [
        "# Garrido Workbook / DES Fidelity Audit",
        "",
        f"Generated: {payload['generated_at']}",
        "",
        "## Workbook Roles",
        "",
    ]
    for item in payload["workbook_roles"]:
        lines.append(f"- `{item['workbook']}`: {item['role']}")
    lines.extend(
        [
            "",
            "## Extraction Gate",
            "",
            f"- Raw rows audited: `{raw_totals['total_rows']}`",
            f"- Recomputed ReT mismatches: `{raw_totals['total_mismatches']}`",
            f"- Max formula gap: `{raw_totals['max_abs_diff']}`",
            f"- CF2 present: `{'CF2' in raw_totals['sheets']}`",
            "",
            "## Main Finding",
            "",
            "The raw workbooks are trajectory/order-level targets. They verify the Excel formula exactly, but the raw per-order mean is dominated by branch composition: risk-active orders use APj/LT or 0.5/RPj, while risk-inactive orders use the running fill-rate branch. Rsult_1 is an aggregate/distribution workbook and should be used as secondary validation, not as a one-to-one replay target.",
            "",
            "## DES Divergence Signals",
            "",
        ]
    )
    if payload.get("des_branch_aggregate"):
        for row in payload["des_branch_aggregate"]:
            if row["policy"] in ("original_S1_I0", "I168_S1") and row["regime"] == "current":
                lines.append(
                    f"- `{row['policy']}` current: fill `{row['fill_rate_order_level_mean']:.3f}`, "
                    f"Excel ReT `{row['mean_ret_excel_formula_mean']:.3f}`, "
                    f"risk-inactive share `{row['risk_inactive_share_mean']:.3f}`, "
                    f"RPj median+ `{row['RPj_median_positive_mean']:.1f} h`, "
                    f"RPj p95+ `{row['RPj_p95_positive_mean']:.1f} h`."
                )
    lines.extend(
        [
            "",
            "Interpretation: if the DES has many risk-inactive orders or much longer positive RPj than the raw Excel targets, raw ReT means can diverge even when service improves. That is a fidelity problem to track separately from the final service/resource optimization panel.",
        ]
    )
    return "\n".join(lines) + "\n"


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--raw-workbooks", nargs="+", type=Path, default=list(DEFAULT_RAW_WORKBOOKS))
    parser.add_argument("--rsult-workbook", type=Path, default=DEFAULT_RSULT_WORKBOOK)
    parser.add_argument("--output-dir", type=Path, default=Path("outputs/audits/garrido_workbook_fidelity_2026-06-26"))
    parser.add_argument("--des-seeds", default="1,2")
    parser.add_argument("--des-regimes", default="current")
    parser.add_argument(
        "--demand-on-hand-fulfillment-delay",
        type=float,
        default=P["demand_on_hand_fulfillment_delay"],
    )
    parser.add_argument("--skip-des", action="store_true")
    args = parser.parse_args()

    out = args.output_dir
    out.mkdir(parents=True, exist_ok=True)
    raw_paths = [Path(path) for path in args.raw_workbooks]
    all_paths = raw_paths + [Path(args.rsult_workbook)]
    targets = load_raw_garrido_targets(raw_paths)
    raw_rows = raw_cf_rows(raw_paths)
    inventory_rows = workbook_inventory(all_paths)
    rsult = rsult_rows(Path(args.rsult_workbook))
    des_rows: list[dict[str, Any]] = []
    des_agg: list[dict[str, Any]] = []
    if not args.skip_des:
        seeds = [int(part.strip()) for part in args.des_seeds.split(",") if part.strip()]
        regimes = [part.strip() for part in args.des_regimes.split(",") if part.strip()]
        des_rows = des_branch_rows(
            seeds,
            regimes,
            demand_on_hand_fulfillment_delay=float(
                args.demand_on_hand_fulfillment_delay
            ),
        )
        des_agg = aggregate_des_rows(des_rows)

    payload = {
        "generated_at": __import__("datetime").datetime.now(__import__("datetime").timezone.utc).isoformat(),
        "workbook_roles": [
            {"workbook": Path(path).name, "path": str(path), "role": RAW_ROLES.get(Path(path).name, "")}
            for path in all_paths
        ],
        "raw_formula_audit": audit_raw_garrido_formula(targets),
        "workbook_inventory": inventory_rows,
        "raw_cf_summary": raw_rows,
        "rsult_summary": rsult,
        "des_branch_rows": des_rows,
        "des_branch_aggregate": des_agg,
        "artifacts": {
            "summary_json": str(out / "audit_summary.json"),
            "report_md": str(out / "audit_report.md"),
            "workbook_inventory_csv": str(out / "workbook_inventory.csv"),
            "raw_cf_summary_csv": str(out / "raw_cf_summary.csv"),
            "rsult_summary_csv": str(out / "rsult_summary.csv"),
            "des_branch_rows_csv": str(out / "des_branch_rows.csv"),
            "des_branch_aggregate_csv": str(out / "des_branch_aggregate.csv"),
        },
    }
    write_csv(out / "workbook_inventory.csv", inventory_rows)
    write_csv(out / "raw_cf_summary.csv", raw_rows)
    write_csv(out / "rsult_summary.csv", rsult)
    write_csv(out / "des_branch_rows.csv", des_rows)
    write_csv(out / "des_branch_aggregate.csv", des_agg)
    (out / "audit_summary.json").write_text(json.dumps(payload, indent=2), encoding="utf-8")
    (out / "audit_report.md").write_text(build_markdown(payload), encoding="utf-8")
    print(f"Wrote {out / 'audit_report.md'}")
    print(f"Raw rows audited: {payload['raw_formula_audit']['total_rows']}")
    print(f"Formula mismatches: {payload['raw_formula_audit']['total_mismatches']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
