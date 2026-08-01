#!/usr/bin/env python3
"""Audit the WRAP/Garrido source bundle without modifying any source artifact.

The audit is intentionally narrower than a behavioral replication.  It records source
fingerprints, workbook coverage, the published Cf1-Cf90 design, and the current claim
boundary.  It is the machine-readable companion to
``docs/GARRIDO_WRAP_SCRES_AI_CONTRACT_V1.md``.
"""
from __future__ import annotations

import argparse
from datetime import datetime, timezone
import hashlib
import json
from pathlib import Path
import re
import sys
from typing import Any, Mapping

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))


def _home_path(*parts: str) -> Path:
    return Path.home().joinpath(*parts)


DEFAULT_SOURCE_CANDIDATES: dict[str, tuple[Path, ...]] = {
    "garrido_2024_scres_ai": (
        _home_path(
            "Library",
            "CloudStorage",
            "GoogleDrive-chisicathomas@gmail.com",
            "My Drive",
            "Supernote",
            "Document",
            "01_RESEARCH",
            "PhD-Papers",
            "garrido2024 scres+AI.pdf",
        ),
        _home_path("Downloads", "garrido2024 scres+AI.pdf"),
    ),
    "garrido_2024_factory_resilience": (
        _home_path("Downloads", "garrido et al 2024 factory resilience.pdf"),
    ),
    "v0_pdf": (_home_path("Downloads", "v.0_neuralNet-scres.pdf"),),
    "v0_docx": (_home_path("Downloads", "v.0_neuralNet-scres.docx"),),
    "raw_data1": (_home_path("Downloads", "Raw_data1+Re.xlsx"),),
    "raw_data2": (_home_path("Downloads", "Raw_data2+Re.xlsx"),),
    "rsult_1": (_home_path("Downloads", "Rsult_1.xlsx"),),
    "wrap_thesis": (
        _home_path(
            "Library",
            "CloudStorage",
            "GoogleDrive-chisicathomas@gmail.com",
            "My Drive",
            "Archive",
            "Misc_Unsorted",
            "Unsorted",
            "WRAP_Theses_Garrido_Rios_2017.pdf",
        ),
    ),
}

WORKBOOK_KEYS = ("raw_data1", "raw_data2", "rsult_1")
CF_PATTERN = re.compile(r"^cf(\d+)$", re.IGNORECASE)


def sha256_file(path: Path) -> str | None:
    if not path.exists() or not path.is_file():
        return None
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def resolve_default_source(key: str) -> Path:
    for candidate in DEFAULT_SOURCE_CANDIDATES[key]:
        if candidate.exists():
            return candidate
    return DEFAULT_SOURCE_CANDIDATES[key][0]


def source_record(key: str, path: Path) -> dict[str, Any]:
    exists = path.exists() and path.is_file()
    return {
        "key": key,
        "path": str(path),
        "exists": exists,
        "bytes": path.stat().st_size if exists else None,
        "sha256": sha256_file(path),
    }


def _header_preview(ws: Any) -> list[list[Any]]:
    max_row = min(int(ws.max_row or 0), 3)
    max_column = min(int(ws.max_column or 0), 8)
    if max_row == 0 or max_column == 0:
        return []
    return [
        list(row)
        for row in ws.iter_rows(
            min_row=1,
            max_row=max_row,
            min_col=1,
            max_col=max_column,
            values_only=True,
        )
    ]


def workbook_record(path: Path) -> dict[str, Any]:
    record: dict[str, Any] = {
        "path": str(path),
        "exists": path.exists() and path.is_file(),
        "cf_sheets": [],
        "all_sheets": [],
        "sheet_inventory": {},
    }
    if not record["exists"]:
        return record

    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        record["all_sheets"] = list(workbook.sheetnames)
        cf_sheets: list[tuple[int, str]] = []
        for sheet_name in workbook.sheetnames:
            match = CF_PATTERN.match(sheet_name)
            if match:
                cf_sheets.append((int(match.group(1)), sheet_name))
        cf_sheets.sort()
        record["cf_sheets"] = [number for number, _ in cf_sheets]
        for number, sheet_name in cf_sheets:
            ws = workbook[sheet_name]
            record["sheet_inventory"][sheet_name] = {
                "used_rows": int(ws.max_row or 0),
                "used_columns": int(ws.max_column or 0),
                "header_preview": _header_preview(ws),
            }
    finally:
        workbook.close()
    return record


def repo_artifact_status(root: Path) -> dict[str, Any]:
    paths = {
        "reproduction": root / "results/garrido_reproduction/reproduction.json",
        "drivers": root / "results/garrido_drivers_per_configuration/result.json",
        "fig5": root / "results/garrido_fig5_surrogate/result.json",
    }
    result: dict[str, Any] = {}
    for name, path in paths.items():
        item: dict[str, Any] = {"path": str(path), "exists": path.exists()}
        if path.exists():
            try:
                payload = json.loads(path.read_text())
                item["claim_status"] = payload.get("claim_status")
                item["schema_version"] = payload.get("schema_version")
            except (OSError, json.JSONDecodeError) as exc:
                item["error"] = f"{type(exc).__name__}: {exc}"
        result[name] = item
    return result


def build_audit(
    *,
    root: Path = ROOT,
    source_paths: Mapping[str, Path] | None = None,
) -> dict[str, Any]:
    """Build an audit payload without changing any input file."""
    paths = {
        key: Path(value)
        for key, value in (source_paths or {}).items()
    }
    for key in DEFAULT_SOURCE_CANDIDATES:
        paths.setdefault(key, resolve_default_source(key))

    sources = {
        key: source_record(key, path)
        for key, path in sorted(paths.items())
    }
    workbooks = {
        key: workbook_record(paths[key])
        for key in WORKBOOK_KEYS
    }

    from supply_chain.config import (
        BACKORDER_QUEUE_CAP,
        HOURS_PER_YEAR_THESIS,
        RET_RE_MIN,
        SIMULATION_HORIZON,
        THESIS_FAITHFUL_PROTOCOL,
    )
    from supply_chain.garrido_thesis_design import DESIGN, VALIDATABLE

    expected_cf = list(range(1, 91))
    author_cf = sorted(
        {
            cf
            for key in ("raw_data1", "raw_data2")
            for cf in workbooks[key]["cf_sheets"]
        }
    )
    generated_cf = [cf for cf in expected_cf if cf not in author_cf]
    all_present = all(item["exists"] for item in sources.values())
    author_cf_matches_expected = author_cf == list(VALIDATABLE)

    payload: dict[str, Any] = {
        "schema_version": "garrido_wrap_source_audit_v1",
        "contract_id": "garrido_wrap_scres_ai_v1",
        "claim_status": (
            "DEVELOPMENT_SOURCE_AUDIT" if all_present else "HOLD_SOURCE_MISSING"
        ),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "sources": sources,
        "workbooks": workbooks,
        "cf_coverage": {
            "published_design": expected_cf,
            "repo_design_count": len(DESIGN),
            "author_delivered_cf": author_cf,
            "author_delivered_matches_cf1_cf20": author_cf_matches_expected,
            "regenerated_from_thesis_design": generated_cf,
            "validatable_constant": list(VALIDATABLE),
            "rsult_cf_sheets": workbooks["rsult_1"]["cf_sheets"],
        },
        "physical_protocol": {
            "protocol": THESIS_FAITHFUL_PROTOCOL["protocol"],
            "year_basis": THESIS_FAITHFUL_PROTOCOL["year_basis"],
            "hours_per_year": HOURS_PER_YEAR_THESIS,
            "horizon_hours": SIMULATION_HORIZON,
            "warmup_trigger": THESIS_FAITHFUL_PROTOCOL["warmup_trigger"],
            "rl_enabled": THESIS_FAITHFUL_PROTOCOL["rl_enabled"],
            "priming_enabled": THESIS_FAITHFUL_PROTOCOL["priming_enabled"],
            "action_multipliers_enabled": THESIS_FAITHFUL_PROTOCOL[
                "action_multipliers_enabled"
            ],
        },
        "metric_contract": {
            "ret_endpoint_status": "PROVISIONAL_UNTIL_SUM_BT_SEMANTICS_CLOSED",
            "sumBt_status": "UNRESOLVED_SOURCE_SEMANTICS",
            "repo_backorder_cap": BACKORDER_QUEUE_CAP,
            "Re_min": RET_RE_MIN,
            "secondary_endpoints_required": [
                "fill_rate",
                "flow_fill_rate",
                "backorder_qty_final",
                "service_loss_auc_ration_hours",
            ],
        },
        "gate_status": {
            "source_manifest": "PASS_SOURCE_AUDIT" if all_present else "HOLD",
            "cf_coverage": (
                "PASS_SOURCE_AUDIT"
                if author_cf_matches_expected and len(DESIGN) == 90
                else "HOLD"
            ),
            "metric_semantics": "HOLD_METRIC_PROVISIONAL",
            "behavioral_fidelity": "HOLD_WRAP_BEHAVIORAL_FIDELITY",
        },
        "repo_artifacts": repo_artifact_status(root),
        "falsifiers": {
            "all_expected_cf_design_rows_exist": len(DESIGN) == 90,
            "author_workbooks_are_not_called_full_cf1_cf90": author_cf_matches_expected,
            "missing_source_files_are_visible": True,
        },
    }
    return payload


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output",
        type=Path,
        default=ROOT / "results/garrido_wrap_source_audit/result.json",
    )
    parser.add_argument("--root", type=Path, default=ROOT)
    for key, option in (
        ("garrido_2024_scres_ai", "--garrido-scres-ai"),
        ("garrido_2024_factory_resilience", "--garrido-factory"),
        ("v0_pdf", "--v0-pdf"),
        ("v0_docx", "--v0-docx"),
        ("raw_data1", "--raw-data1"),
        ("raw_data2", "--raw-data2"),
        ("rsult_1", "--rsult-1"),
        ("wrap_thesis", "--wrap-thesis"),
    ):
        parser.add_argument(option, dest=key, type=Path)
    return parser


def main() -> int:
    args = build_parser().parse_args()
    overrides = {
        key: getattr(args, key)
        for key in DEFAULT_SOURCE_CANDIDATES
        if getattr(args, key) is not None
    }
    payload = build_audit(root=args.root, source_paths=overrides)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n")
    print(f"Saved: {args.output}")
    print(f"Claim status: {payload['claim_status']}")
    print(
        "Author workbook Cf coverage: "
        f"{payload['cf_coverage']['author_delivered_cf']}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
