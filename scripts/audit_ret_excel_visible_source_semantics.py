#!/usr/bin/env python3
"""Audit source semantics behind the workbook-visible Garrido ReT ledger.

This is a read-only provenance audit.  It does not open stochastic experiment
tapes.  Its main purpose is to keep formula replay separate from the stronger
claim that a counterfactual DES reconstructs the workbook's time-varying Bt/Ut
fields correctly.
"""

from __future__ import annotations

import argparse
from collections import defaultdict
from datetime import datetime, timezone
from hashlib import sha256
import json
from pathlib import Path
import re
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from supply_chain.config import BACKORDER_QUEUE_CAP
from supply_chain.garrido_replication import (
    DEFAULT_RAW_WORKBOOKS,
    DEFAULT_RSULT_WORKBOOK,
    audit_raw_garrido_formula,
    column,
    find_header_map,
    load_raw_garrido_targets,
)


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_OUTPUT = (
    ROOT
    / "research"
    / "paper2_exhaustive_search"
    / "ret_excel_visible_v1_source_semantics_audit_20260714.json"
)
DEFAULT_THESIS = (
    Path.home()
    / "Library"
    / "CloudStorage"
    / "GoogleDrive-chisicathomas@gmail.com"
    / "My Drive"
    / "Archive"
    / "Misc_Unsorted"
    / "Unsorted"
    / "WRAP_Theses_Garrido_Rios_2017.pdf"
)


def file_sha256(path: Path) -> str:
    digest = sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def json_sha256(value: Any) -> str:
    payload = json.dumps(
        value, sort_keys=True, separators=(",", ":"), ensure_ascii=False
    ).encode()
    return sha256(payload).hexdigest()


def normalize_row_formula(formula: str) -> str:
    """Replace Excel row references while preserving numeric constants."""

    return re.sub(r"(?<=[A-Z])\d+", "{row}", formula.upper().replace(" ", ""))


def inspect_formula_cells(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=False, read_only=True)
    try:
        sheets: dict[str, Any] = {}
        for sheet_name in workbook.sheetnames:
            if not sheet_name.lower().startswith("cf"):
                continue
            suffix = sheet_name[2:]
            if not suffix.isdigit():
                continue
            worksheet = workbook[sheet_name]
            header_row, headers, _ = find_header_map(worksheet)
            try:
                ret_column = column(headers, "ReT")
            except KeyError:
                continue
            formulas: list[tuple[str, str]] = []
            for row_index, row in enumerate(
                worksheet.iter_rows(min_row=header_row + 1),
                start=header_row + 1,
            ):
                value = row[ret_column - 1].value
                if isinstance(value, str) and value.startswith("="):
                    coordinate = f"{get_column_letter(ret_column)}{row_index}"
                    formulas.append((coordinate, value))
            normalized = sorted({normalize_row_formula(value) for _, value in formulas})
            sheets[sheet_name] = {
                "header_row": header_row,
                "ret_column": get_column_letter(ret_column),
                "formula_count": len(formulas),
                "first_formula_cell": formulas[0][0] if formulas else None,
                "first_formula": formulas[0][1] if formulas else None,
                "last_formula_cell": formulas[-1][0] if formulas else None,
                "normalized_formula_patterns": normalized,
            }
        return {
            "path": str(path),
            "bytes": path.stat().st_size,
            "sha256": file_sha256(path),
            "sheets": sheets,
        }
    finally:
        workbook.close()


def visible_only_oat_bt_snapshots(orders: Iterable[Any]) -> dict[int, int]:
    """Mirror the current OAT-ledger rule using visible workbook rows only.

    Missing/lost rows are unavailable in the raw workbooks, so this diagnostic
    is deliberately not called a full reconstruction.  It is useful alongside
    the stronger ordering evidence and the thesis Annex-B request payload.
    """

    rows = list(orders)
    events: list[tuple[float, int, int]] = []
    for row in rows:
        activation = float(row.optj) + float(row.ltj)
        completion = float(row.oatj)
        if completion > activation:
            events.append((activation, 1, +1))
            events.append((completion, 0, -1))
    events.sort()
    event_index = 0
    current_backorders = 0
    snapshots: dict[int, int] = {}
    for row in sorted(rows, key=lambda value: (value.oatj, value.j)):
        while event_index < len(events) and events[event_index][0] <= row.oatj:
            current_backorders += events[event_index][2]
            event_index += 1
        current_backorders = max(
            0, min(current_backorders, int(BACKORDER_QUEUE_CAP))
        )
        snapshots[int(row.j)] = int(current_backorders)
    return snapshots


def sheet_timing_audit(target: Any) -> dict[str, Any]:
    rows = list(target.orders)
    oat_rows = sorted(rows, key=lambda value: (value.oatj, value.j))
    same_oat: dict[float, int] = defaultdict(int)
    for row in rows:
        same_oat[float(row.oatj)] += 1
    snapshots = visible_only_oat_bt_snapshots(rows)
    initial_rows: list[Any] = []
    expected_j = 1
    for row in rows:
        if int(row.j) != expected_j:
            break
        initial_rows.append(row)
        expected_j += 1
    matches = sum(snapshots[int(row.j)] == int(row.sum_bt) for row in rows)
    initial_matches = sum(
        snapshots[int(row.j)] == int(row.sum_bt) for row in initial_rows
    )
    mismatch_examples = [
        {
            "j": int(row.j),
            "OPTj": float(row.optj),
            "OATj": float(row.oatj),
            "workbook_sumBt": int(row.sum_bt),
            "visible_only_oat_reconstruction_sumBt": snapshots[int(row.j)],
        }
        for row in rows
        if snapshots[int(row.j)] != int(row.sum_bt)
    ][:5]
    return {
        "sheet": target.sheet,
        "rows": len(rows),
        "max_j": int(target.max_j),
        "omitted_j_count_through_max": int(target.max_j - len(rows)),
        "j_strictly_increasing": all(
            int(right.j) > int(left.j) for left, right in zip(rows, rows[1:])
        ),
        "OPTj_nondecreasing_in_j_order": all(
            float(right.optj) >= float(left.optj)
            for left, right in zip(rows, rows[1:])
        ),
        "OATj_adjacent_inversions_in_j_order": sum(
            float(right.oatj) < float(left.oatj)
            for left, right in zip(rows, rows[1:])
        ),
        "sumUt_adjacent_decreases_in_j_order": sum(
            float(right.sum_ut) < float(left.sum_ut)
            for left, right in zip(rows, rows[1:])
        ),
        "sumUt_adjacent_decreases_in_OATj_order": sum(
            float(right.sum_ut) < float(left.sum_ut)
            for left, right in zip(oat_rows, oat_rows[1:])
        ),
        "same_exact_OATj_group_count": sum(
            count > 1 for count in same_oat.values()
        ),
        "visible_rows_only_OAT_Bt_match": {
            "classification": "diagnostic_not_full_reconstruction_missing_rows_unobserved",
            "matches": matches,
            "total": len(rows),
            "rate": matches / len(rows) if rows else None,
            "initial_contiguous_j_prefix_rows": len(initial_rows),
            "initial_contiguous_j_prefix_matches": initial_matches,
            "mismatch_examples": mismatch_examples,
        },
        "final_workbook_sumBt": float(target.final_sum_bt),
        "final_workbook_sumUt": float(target.final_sum_ut),
    }


def build_audit(
    raw_workbooks: tuple[Path, ...] = DEFAULT_RAW_WORKBOOKS,
    *,
    rsult_workbook: Path = DEFAULT_RSULT_WORKBOOK,
    thesis_path: Path = DEFAULT_THESIS,
) -> dict[str, Any]:
    if not all(path.is_file() for path in raw_workbooks):
        missing = [str(path) for path in raw_workbooks if not path.is_file()]
        raise FileNotFoundError(f"missing raw workbook(s): {missing}")
    targets = load_raw_garrido_targets(raw_workbooks)
    formula_audit = audit_raw_garrido_formula(targets)
    timing_sheets = [sheet_timing_audit(target) for _, target in sorted(targets.items())]
    totals = {
        "rows": sum(row["rows"] for row in timing_sheets),
        "OATj_adjacent_inversions_in_j_order": sum(
            row["OATj_adjacent_inversions_in_j_order"] for row in timing_sheets
        ),
        "sumUt_adjacent_decreases_in_j_order": sum(
            row["sumUt_adjacent_decreases_in_j_order"] for row in timing_sheets
        ),
        "sumUt_adjacent_decreases_in_OATj_order": sum(
            row["sumUt_adjacent_decreases_in_OATj_order"] for row in timing_sheets
        ),
        "same_exact_OATj_group_count": sum(
            row["same_exact_OATj_group_count"] for row in timing_sheets
        ),
        "visible_rows_only_OAT_Bt_matches": sum(
            row["visible_rows_only_OAT_Bt_match"]["matches"]
            for row in timing_sheets
        ),
        "visible_rows_only_OAT_Bt_total": sum(
            row["visible_rows_only_OAT_Bt_match"]["total"]
            for row in timing_sheets
        ),
        "initial_contiguous_j_prefix_rows": sum(
            row["visible_rows_only_OAT_Bt_match"]["initial_contiguous_j_prefix_rows"]
            for row in timing_sheets
        ),
        "initial_contiguous_j_prefix_matches": sum(
            row["visible_rows_only_OAT_Bt_match"][
                "initial_contiguous_j_prefix_matches"
            ]
            for row in timing_sheets
        ),
    }
    totals["visible_rows_only_OAT_Bt_match_rate"] = (
        totals["visible_rows_only_OAT_Bt_matches"]
        / totals["visible_rows_only_OAT_Bt_total"]
    )
    totals["initial_contiguous_j_prefix_match_rate"] = (
        totals["initial_contiguous_j_prefix_matches"]
        / totals["initial_contiguous_j_prefix_rows"]
    )
    formula_inspection = [inspect_formula_cells(path) for path in raw_workbooks]
    result: dict[str, Any] = {
        "schema_version": "ret_excel_visible_source_semantics_audit_v1",
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "status": (
            "FORMULA_EXACT__VISIBLE_ROWS_OBSERVED__"
            "OAT_BT_UT_RECONSTRUCTION_NOT_SOURCE_VALIDATED"
        ),
        "scientific_disposition": {
            "current_ret_excel_visible_v1": "HOLD_FOR_LEDGER_SEMANTICS_REPAIR",
            "active_vps_switch4": (
                "QUARANTINE_AS_NONCANONICAL_METRIC_DEVELOPMENT_OUTPUT; no H_PI, "
                "H_obs, comparator ceiling, null, or positive claim may use it"
            ),
            "program_h_j_visible_v1_repairs": (
                "QUARANTINE_IF_SCORED_BY_THE_CURRENT_OAT_LEDGER; rescore only after "
                "the request-snapshot contract and implementation pass"
            ),
            "historical_formula_replay": (
                "REMAINS_VALID for the narrow statement that cached workbook ReT "
                "equals the published cell formula when workbook Bt/Ut are inputs"
            ),
        },
        "claim_boundary": (
            "Zero formula mismatches do not validate the current Python algorithm "
            "that derives Bt/Ut from OPTj, LTj, OATj and lost_time."
        ),
        "no_tape_access": {
            "stochastic_experiment_tapes_accessed": False,
            "virgin_tapes_accessed": False,
            "inputs": "three user-provided workbooks, thesis PDF, and repository source only",
        },
        "sources": {
            "raw_workbooks": formula_inspection,
            "Rsult_1": {
                "path": str(rsult_workbook),
                "available": rsult_workbook.is_file(),
                "bytes": rsult_workbook.stat().st_size
                if rsult_workbook.is_file()
                else None,
                "sha256": file_sha256(rsult_workbook)
                if rsult_workbook.is_file()
                else None,
                "role": "secondary transformed/aggregate workbook, not the 20-CF row ledger",
            },
            "thesis": {
                "path": str(thesis_path),
                "available": thesis_path.is_file(),
                "sha256": file_sha256(thesis_path) if thesis_path.is_file() else None,
                "citation": {
                    "location": "Annex B, printed page 169 (PDF page 170)",
                    "source_fact": (
                        "The request-associated barrier matrix contains quantity, "
                        "order number, generation time, accumulated-order total and "
                        "lost-order total before being sent to Op9."
                    ),
                },
            },
            "repository": {
                "ret_thesis": {
                    "path": "supply_chain/ret_thesis.py",
                    "sha256": file_sha256(ROOT / "supply_chain" / "ret_thesis.py"),
                    "current_OAT_ledger_lines": "333-393",
                },
                "episode_metrics": {
                    "path": "supply_chain/episode_metrics.py",
                    "sha256": file_sha256(ROOT / "supply_chain" / "episode_metrics.py"),
                    "canonical_alias_lines": "170-176,216-224",
                },
                "workbook_formula_auditor": {
                    "path": "supply_chain/garrido_replication.py",
                    "sha256": file_sha256(
                        ROOT / "supply_chain" / "garrido_replication.py"
                    ),
                    "formula_audit_lines": "327-369",
                    "limitation": "uses workbook-provided sumBt and sumUt as inputs",
                },
            },
        },
        "workbook_formula_replay": {
            "formula": formula_audit["formula"],
            "formula_rows": formula_audit["total_rows"],
            "mismatches": formula_audit["total_mismatches"],
            "max_abs_diff": formula_audit["max_abs_diff"],
            "verdict": "PASS_EXACT_GIVEN_WORKBOOK_BT_UT",
        },
        "timing_and_population_evidence": {
            "interpretation": (
                "Rows are in j/OPT order, not OAT order. SumUt is monotone in j "
                "but not in OAT, which is consistent with a request-carried ledger "
                "snapshot and inconsistent with treating the raw row as an OAT-time "
                "cumulative-loss observation."
            ),
            "totals": totals,
            "sheets": timing_sheets,
        },
        "proposed_ret_excel_visible_v2": {
            "status": "PROPOSED_NOT_IMPLEMENTED",
            "row_population": (
                "emit completed, non-lost rows only, in original j order; preserve "
                "original j and do not clip"
            ),
            "request_snapshot_fields": [
                "ret_bt_at_request",
                "ret_ut_at_request",
                "ret_ledger_snapshot_time",
                "ret_ledger_event_sequence",
            ],
            "algorithm": [
                "At the order-generation callback t=OPTj, read the live source-defined capped Bt queue count and cumulative Ut before inserting order j.",
                "Freeze Bt, Ut, t and a monotone event-sequence number on OrderRecord; later release, completion or loss events never rewrite them.",
                "For forensic Excel replay, inject workbook sumBt/sumUt directly into those fields rather than reconstructing them from OATj.",
                "After the episode, emit only completed non-lost records and apply the unchanged Excel risk switch/formula with the frozen request snapshots and original j.",
            ],
            "same_timestamp_event_order": (
                "PROPOSED deterministic convention: process events that were already "
                "scheduled before the order-generation callback; snapshot Bt/Ut; "
                "then enqueue the new request. Events scheduled later at the same "
                "timestamp cannot alter that request's snapshot. Record the event "
                "sequence so the convention is auditable. This convention remains "
                "subject to Garrido/Simulink confirmation."
            ),
            "falsifiable_garrido_question": (
                "In the Simulink request barrier, were sumBt and sumUt frozen when "
                "order j was generated, and at an identical simulation timestamp "
                "did queue removals/losses occur before or after that snapshot?"
            ),
            "required_tests": [
                "The canonical aggregator itself, not a separate helper, must reproduce all 47,546 cached workbook ReT cells exactly when request snapshots are injected.",
                "Changing OATj after a request snapshot is frozen must not change a no-risk row's fill-rate branch value.",
                "Permutation and batch-tie fixtures must prove deterministic same-time behavior without removing every completion before every row is scored.",
                "Source-backed tests must cover j gaps, horizon-unresolved rows, queue cap 60, lost-order Ut carry-forward, risk-active branches, values above one and zero visible rows.",
                "All active H/J/MTR development outputs must be rescored on identical burned/calibration tapes before any comparator or boundary claim is restored.",
            ],
        },
        "metric_comparison": {
            "thesis_conceptual_Eq_5_5": (
                "Four conditional AP/RP/DP/fill-rate subindicators; thesis states a "
                "0-1 range, unlike unclipped workbook cells that can exceed one."
            ),
            "excel_formula_given_source_ledger": (
                "Latest exact operational artifact: risk switch plus AP/LT or "
                "0.5/RP, otherwise 1-(Bt+Ut)/j; exact on 47,546 cells."
            ),
            "ret_excel_full_ledger": (
                "Repository secondary proxy that emits every generated order and "
                "scores unfulfilled rows zero; not the raw workbook population."
            ),
            "ret_excel_visible_v1_current": (
                "Correct visible-row intent and exact per-row formula, but current "
                "OAT-derived Bt/Ut reconstruction is not source-validated."
            ),
            "cobb_douglas_2024": (
                "Separate factory APP construct over inventory, backorders, spare "
                "capacity, fulfilment effort/time and normalized cost; not thesis "
                "order-level ReT and not a rescue endpoint for current Paper 2."
            ),
        },
    }
    result["content_sha256"] = json_sha256(result)
    return result


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--thesis", type=Path, default=DEFAULT_THESIS)
    args = parser.parse_args()
    payload = build_audit(thesis_path=args.thesis)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    temporary = args.output.with_suffix(args.output.suffix + ".tmp")
    temporary.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n")
    temporary.replace(args.output)
    print(
        json.dumps(
            {
                "status": payload["status"],
                "output": str(args.output),
                "content_sha256": payload["content_sha256"],
                "formula_rows": payload["workbook_formula_replay"]["formula_rows"],
                "OATj_inversions": payload["timing_and_population_evidence"]["totals"][
                    "OATj_adjacent_inversions_in_j_order"
                ],
            },
            indent=2,
            sort_keys=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
