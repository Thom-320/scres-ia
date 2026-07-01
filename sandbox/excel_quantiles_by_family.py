#!/usr/bin/env python3
"""Extract Garrido Excel CTj/RPj/DPj/APj quantiles by risk family.

Family CF1-CF10 -> R1 (Raw_data1+Re.xlsx, 5 risk cols).
Family CF11-CF20 -> R2 (Raw_data2+Re.xlsx, 11 risk cols).

Headers are read dynamically (CF2 header is on row 2; Raw_data2 has a wider
risk-column block than Raw_data1). This is the Excel-side target for the
CTj/RPj/DPj fidelity comparison requested by GARRIDO_ORIGINAL_RUNS_GATE.
"""
from __future__ import annotations

import argparse
import json
import warnings
from pathlib import Path

import numpy as np
import openpyxl

warnings.filterwarnings("ignore")

DOWNLOADS = Path.home() / "Downloads"
FILES = {
    "R1": (DOWNLOADS / "Raw_data1+Re.xlsx", [f"CF{i}" for i in range(1, 11)]),
    "R2": (DOWNLOADS / "Raw_data2+Re.xlsx", [f"CF{i}" for i in range(11, 21)]),
}


def _find(hdr, name):
    for i, h in enumerate(hdr):
        if h is not None and str(h).strip().lower() == name.lower():
            return i
    return None


def _header_row(ws):
    for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if _find(row, "CTj") is not None and _find(row, "OPTj") is not None:
            return ri, row
    raise RuntimeError("header row not found")


def quantiles(values):
    arr = np.asarray([v for v in values if isinstance(v, (int, float))], dtype=float)
    arr = arr[np.isfinite(arr)]
    if arr.size == 0:
        return {"n": 0}
    qs = [50, 90, 95, 99]
    out = {"n": int(arr.size), "mean": float(arr.mean())}
    for q in qs:
        out[f"p{q}"] = float(np.percentile(arr, q))
    out["max"] = float(arr.max())
    return out


def extract_family(path, sheets):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    per_cf = {}
    pools = {col: [] for col in ("CTj", "RPj", "DPj", "APj")}
    for s in sheets:
        ws = wb[s]
        _, hdr = _header_row(ws)
        idx = {c: _find(hdr, c) for c in ("CTj", "RPj", "DPj", "APj")}
        if any(v is None for v in idx.values()):
            continue
        start = _header_row(ws)[0]
        cols = {c: [] for c in idx}
        for row in ws.iter_rows(min_row=start + 1, values_only=True):
            for c, i in idx.items():
                v = row[i] if i < len(row) else None
                if isinstance(v, (int, float)):
                    cols[c].append(v)
                    pools[c].append(v)
        per_cf[s] = {c: quantiles(cols[c]) for c in idx}
    wb.close()
    return per_cf, {c: quantiles(pools[c]) for c in pools}


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--output", type=Path,
                    default=Path("sandbox/results/excel_quantiles_by_family.json"))
    args = ap.parse_args()
    args.output.parent.mkdir(parents=True, exist_ok=True)

    result = {}
    for fam, (path, sheets) in FILES.items():
        per_cf, pooled = extract_family(path, sheets)
        result[fam] = {"workbook": path.name, "per_cf": per_cf, "pooled": pooled}
        print(f"\n=== Family {fam} ({path.name}) — pooled order-level quantiles ===")
        for col in ("CTj", "RPj", "DPj", "APj"):
            q = pooled[col]
            if q.get("n", 0):
                print(f"  {col:4} n={q['n']:6d} mean={q['mean']:9.1f} "
                      f"p50={q['p50']:9.1f} p90={q['p90']:9.1f} "
                      f"p95={q['p95']:9.1f} p99={q['p99']:9.1f} max={q['max']:10.1f}")
    args.output.write_text(json.dumps(result, indent=2), encoding="utf-8")
    print(f"\nSaved: {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
