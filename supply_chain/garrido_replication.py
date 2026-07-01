from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from statistics import fmean, median
from typing import Any, Iterable

from openpyxl import load_workbook


DEFAULT_DOWNLOADS_DIR = Path.home() / "Downloads"
DEFAULT_RAW_WORKBOOKS = (
    DEFAULT_DOWNLOADS_DIR / "Raw_data1+Re.xlsx",
    DEFAULT_DOWNLOADS_DIR / "Raw_data2+Re.xlsx",
)
DEFAULT_RSULT_WORKBOOK = DEFAULT_DOWNLOADS_DIR / "Rsult_1.xlsx"

EXCEL_RET_FORMULA = (
    "IF(AVERAGE(risk_cols)>0, IF(APj>0, APj/LT, 0.5*(1/RPj)), "
    "1-((sumBt+sumUt)/j))"
)


@dataclass(frozen=True)
class GarridoOrderTarget:
    source_file: str
    sheet: str
    cfi: int
    row_index: int
    q: float
    j: int
    optj: float
    oatj: float
    ctj: float
    ltj: float
    sum_bt: float
    apj: float
    rpj: float
    dpj: float
    risk_values: dict[str, float]
    sum_ut: float
    op9: float
    ret: float
    delta_ret: float

    @property
    def risk_active(self) -> bool:
        return any(value > 0.0 for value in self.risk_values.values())

    @property
    def excel_case(self) -> str:
        if not self.risk_active:
            return "excel_fill_rate"
        if self.apj > 0.0:
            return "excel_autotomy"
        if self.rpj > 0.0:
            return "excel_recovery"
        return "excel_risk_no_recovery"

    def as_order_tape_row(self) -> dict[str, Any]:
        return {
            "j": int(self.j),
            "OPTj": float(self.optj),
            "Q": float(self.q),
            "contingent": bool(self.q > 2600.0 or self.risk_values.get("R24", 0.0) > 0.0),
            "ret_attribution": {
                "APj": float(self.apj),
                "RPj": float(self.rpj),
                "DPj": float(self.dpj),
                "LTj": float(self.ltj),
                "risk_values": dict(self.risk_values),
            },
        }


@dataclass(frozen=True)
class GarridoCFTarget:
    source_file: str
    sheet: str
    cfi: int
    seed: int
    warmup_hours: float
    header_row: int
    risk_columns: tuple[str, ...]
    orders: tuple[GarridoOrderTarget, ...]

    @property
    def n_orders(self) -> int:
        return len(self.orders)

    @property
    def horizon_hours(self) -> float:
        return max((order.oatj for order in self.orders), default=0.0)

    @property
    def ret_mean_excel(self) -> float:
        return fmean(order.ret for order in self.orders) if self.orders else float("nan")

    @property
    def first_optj(self) -> float:
        return min((order.optj for order in self.orders), default=float("nan"))

    @property
    def last_oatj(self) -> float:
        return max((order.oatj for order in self.orders), default=float("nan"))

    @property
    def max_sum_bt(self) -> float:
        return max((order.sum_bt for order in self.orders), default=0.0)

    @property
    def max_sum_ut(self) -> float:
        return max((order.sum_ut for order in self.orders), default=0.0)

    @property
    def case_counts_excel_formula(self) -> dict[str, int]:
        counts = {
            "excel_fill_rate": 0,
            "excel_autotomy": 0,
            "excel_recovery": 0,
            "excel_risk_no_recovery": 0,
        }
        for order in self.orders:
            counts[order.excel_case] += 1
        return counts

    def order_tape(self) -> list[dict[str, Any]]:
        return [order.as_order_tape_row() for order in self.orders]


def norm_header(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "")


def to_float(value: Any, default: float = 0.0) -> float:
    if value is None:
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def find_header_map(
    ws: Any, *, max_scan_rows: int = 10
) -> tuple[int, dict[str, int], dict[int, Any]]:
    required = {"j", "apj", "rpj", "dpj", "ret"}
    best: tuple[int, dict[str, int], dict[int, Any]] | None = None
    best_score = -1
    for row_index, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), start=1
    ):
        headers: dict[str, int] = {}
        labels: dict[int, Any] = {}
        for column_index, value in enumerate(row, start=1):
            label = norm_header(value)
            if label:
                headers[label] = column_index
                labels[column_index] = value
        score = len(required.intersection(headers))
        if score > best_score:
            best = (row_index, headers, labels)
            best_score = score
        if score == len(required):
            return row_index, headers, labels
    if best is None:
        return 1, {}, {}
    return best


def column(headers: dict[str, int], *names: str) -> int:
    for name in names:
        key = norm_header(name)
        if key in headers:
            return headers[key]
    raise KeyError(f"Missing any of headers: {names}")


def excel_ret_value(
    *,
    j: float,
    lt: float,
    sum_bt: float,
    apj: float,
    rpj: float,
    sum_ut: float,
    risk_values: Iterable[float],
) -> float:
    if any(value > 0.0 for value in risk_values):
        if apj > 0.0:
            return apj / max(lt, 1e-9)
        if rpj > 0.0:
            return 0.5 * (1.0 / rpj)
        return 0.0
    return 1.0 - ((sum_bt + sum_ut) / max(j, 1e-9))


def _sheet_cfi(sheet_name: str) -> int | None:
    if not sheet_name.lower().startswith("cf"):
        return None
    try:
        return int(sheet_name.lower().replace("cf", ""))
    except ValueError:
        return None


def load_raw_garrido_targets(
    workbooks: Iterable[Path] = DEFAULT_RAW_WORKBOOKS,
) -> dict[int, GarridoCFTarget]:
    targets: dict[int, GarridoCFTarget] = {}
    for path in workbooks:
        wb = load_workbook(path, data_only=True, read_only=True)
        try:
            for sheet_name in wb.sheetnames:
                cfi = _sheet_cfi(sheet_name)
                if cfi is None:
                    continue
                ws = wb[sheet_name]
                header_row, headers, header_labels = find_header_map(ws)
                try:
                    q_col = column(headers, "Q")
                    j_col = column(headers, "j")
                    opt_col = column(headers, "OPTj")
                    oat_col = column(headers, "OATj")
                    ct_col = column(headers, "CTj")
                    lt_col = column(headers, "LT")
                    bt_col = column(headers, "\u2211Bt", "sumBt")
                    ap_col = column(headers, "APj")
                    rp_col = column(headers, "RPj")
                    dp_col = column(headers, "DPj")
                    ut_col = column(headers, "\u2211Ut", "sumUt")
                    op9_col = column(headers, "OP9")
                    ret_col = column(headers, "ReT")
                except KeyError:
                    continue
                delta_col = headers.get(norm_header("\u0394ReT")) or headers.get(
                    norm_header("deltaReT")
                )

                risk_cols = list(range(dp_col + 1, ut_col))
                risk_labels = tuple(str(header_labels.get(col, "")) for col in risk_cols)
                orders: list[GarridoOrderTarget] = []
                seed = 0
                warmup = float("nan")
                for row_index, row in enumerate(
                    ws.iter_rows(min_row=header_row + 1), start=header_row + 1
                ):
                    row_label = norm_header(row[0].value) if row else ""
                    if row_label == "seed" and len(row) > 1:
                        seed = int(to_float(row[1].value, default=0.0))
                    elif row_label == "warm-upperiod" and len(row) > 1:
                        warmup = to_float(row[1].value, default=float("nan"))

                    j_value = row[j_col - 1].value
                    ret_value = row[ret_col - 1].value
                    if j_value is None or ret_value is None:
                        continue
                    risk_values = {
                        label: to_float(row[col - 1].value)
                        for label, col in zip(risk_labels, risk_cols, strict=True)
                    }
                    orders.append(
                        GarridoOrderTarget(
                            source_file=str(path),
                            sheet=sheet_name,
                            cfi=cfi,
                            row_index=row_index,
                            q=to_float(row[q_col - 1].value),
                            j=int(to_float(j_value)),
                            optj=to_float(row[opt_col - 1].value),
                            oatj=to_float(row[oat_col - 1].value),
                            ctj=to_float(row[ct_col - 1].value),
                            ltj=to_float(row[lt_col - 1].value),
                            sum_bt=to_float(row[bt_col - 1].value),
                            apj=to_float(row[ap_col - 1].value),
                            rpj=to_float(row[rp_col - 1].value),
                            dpj=to_float(row[dp_col - 1].value),
                            risk_values=risk_values,
                            sum_ut=to_float(row[ut_col - 1].value),
                            op9=to_float(row[op9_col - 1].value),
                            ret=to_float(ret_value),
                            delta_ret=(
                                to_float(row[delta_col - 1].value)
                                if delta_col is not None
                                else 0.0
                            ),
                        )
                    )
                targets[cfi] = GarridoCFTarget(
                    source_file=str(path),
                    sheet=sheet_name,
                    cfi=cfi,
                    seed=seed,
                    warmup_hours=warmup,
                    header_row=header_row,
                    risk_columns=risk_labels,
                    orders=tuple(orders),
                )
        finally:
            wb.close()
    return targets


def audit_raw_garrido_formula(
    targets: dict[int, GarridoCFTarget], *, tolerance: float = 1e-9
) -> dict[str, Any]:
    total_rows = 0
    total_mismatches = 0
    max_abs_diff = 0.0
    sheets: dict[str, Any] = {}
    for cfi, target in sorted(targets.items()):
        sheet_rows = 0
        sheet_mismatches = 0
        sheet_max = 0.0
        for order in target.orders:
            computed = excel_ret_value(
                j=order.j,
                lt=order.ltj,
                sum_bt=order.sum_bt,
                apj=order.apj,
                rpj=order.rpj,
                sum_ut=order.sum_ut,
                risk_values=order.risk_values.values(),
            )
            diff = abs(computed - order.ret)
            sheet_max = max(sheet_max, diff)
            if diff > tolerance:
                sheet_mismatches += 1
            sheet_rows += 1
        sheets[f"CF{cfi}"] = {
            "rows": sheet_rows,
            "mismatches": sheet_mismatches,
            "max_abs_diff": sheet_max,
            "risk_columns": list(target.risk_columns),
            "header_row": target.header_row,
        }
        total_rows += sheet_rows
        total_mismatches += sheet_mismatches
        max_abs_diff = max(max_abs_diff, sheet_max)
    return {
        "formula": EXCEL_RET_FORMULA,
        "total_rows": total_rows,
        "total_mismatches": total_mismatches,
        "max_abs_diff": max_abs_diff,
        "sheets": sheets,
    }


def summarize_rsult_workbook(path: Path = DEFAULT_RSULT_WORKBOOK) -> dict[str, Any]:
    if not path.exists():
        return {"path": str(path), "available": False, "sheets": {}}

    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        sheets: dict[str, Any] = {}
        for sheet_name in ("APj", "RPj", "DPj", "Re"):
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            sheet_summary: dict[str, Any] = {}
            for col_index, header in enumerate(headers, start=1):
                if header is None:
                    continue
                values = [
                    to_float(row[0].value)
                    for row in ws.iter_rows(
                        min_row=3,
                        min_col=col_index,
                        max_col=col_index,
                    )
                    if row[0].value is not None
                ]
                if not values:
                    continue
                sheet_summary[str(header)] = {
                    "n": len(values),
                    "mean": fmean(values),
                    "median": median(values),
                    "min": min(values),
                    "max": max(values),
                }
            sheets[sheet_name] = sheet_summary
        return {"path": str(path), "available": True, "sheets": sheets}
    finally:
        wb.close()


def target_to_summary(target: GarridoCFTarget) -> dict[str, Any]:
    return {
        "cfi": target.cfi,
        "sheet": target.sheet,
        "workbook": target.source_file,
        "seed": target.seed,
        "warmup_hours": target.warmup_hours,
        "horizon_hours": target.horizon_hours,
        "n_orders": target.n_orders,
        "ret_mean_excel": target.ret_mean_excel,
        "first_optj": target.first_optj,
        "last_oatj": target.last_oatj,
        "max_sum_bt": target.max_sum_bt,
        "max_sum_ut": target.max_sum_ut,
        "risk_columns": list(target.risk_columns),
        "case_counts_excel_formula": target.case_counts_excel_formula,
    }
